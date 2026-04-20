import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import customtkinter as ctk
from datetime import date, datetime
import csv
from connect_db import execute_query

from components.date_picker import make_date_entry

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _XLSX = True
except ImportError:
    _XLSX = False

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    _PDF = True
except ImportError:
    _PDF = False

MAROON     = "#5E1219"
GOLD       = "#EFBF04"
BLACK      = "#000000"
WHITE      = "#FFFFFF"
LIGHT      = "#F9F9F9"
GOLD_LIGHT = "#FFF8E7"

REPORT_TYPES = [
    "Room Usage Report",
    "Student Usage & Penalty Report",
    "Violations Report",
]

FILENAME_MAP = {
    "Room Usage Report":              "RoomUsageReport",
    "Student Usage & Penalty Report": "StudentPenaltyReport",
    "Violations Report":              "ViolationsReport",
}

HEADERS_MAP = {
    "Room Usage Report": [
        "Room", "Category", "Total", "Check-Ins",
        "Cancellations", "No-Shows", "Reserved", "Completed",
    ],
    "Student Usage & Penalty Report": [
        "User ID", "First Name", "Last Name", "Email",
        "Penalty Points", "Status", "Total Reservations", "No-Shows", "Cancellations",
    ],
    "Violations Report": [
        "Violation ID", "First Name", "Last Name", "Room",
        "Type", "Points", "Status", "Date",
    ],
}

ROW_KEYS_MAP = {
    "Room Usage Report": [
        "room_number", "room_category", "total_reservations", "checkins",
        "cancellations", "no_shows", "reserved", "completed",
    ],
    "Student Usage & Penalty Report": [
        "user_id", "first_name", "last_name", "email",
        "penalty_points", "account_status", "total_reservations", "no_shows", "cancellations",
    ],
    "Violations Report": [
        "violation_id", "first_name", "last_name", "room_number",
        "violation_type", "points_assessed", "status", "created_at",
    ],
}


class Reports(tk.Frame):
    def __init__(self, parent, user_info: dict, navigator, **kwargs):
        super().__init__(parent, bg=WHITE, **kwargs)
        self.user_info = user_info
        self.navigator = navigator
        self.current_report_data = []
        self.current_report_type = ""
        self.current_date_range  = ("", "")
        self._build()

    # ------------------------------------------------------------------
    def _build(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        content = tk.Frame(self, bg=WHITE)
        content.grid(row=0, column=0, sticky="nsew", padx=24, pady=16)
        content.columnconfigure(0, weight=1)
        content.rowconfigure(3, weight=1)

        tk.Label(content, text="Generate Reports",
                 fg=MAROON, bg=WHITE, font=("Poppins", 22, "bold")
                 ).grid(row=0, column=0, sticky="w")

        ctrl = tk.Frame(content, bg=WHITE)
        ctrl.grid(row=1, column=0, sticky="ew", pady=(14, 8))

        tk.Label(ctrl, text="Report Type:", fg=BLACK, bg=WHITE,
                 font=("Poppins", 12, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.rtype_var = tk.StringVar(value=REPORT_TYPES[0])
        ttk.Combobox(ctrl, textvariable=self.rtype_var,
                     values=REPORT_TYPES, state="readonly", width=34,
                     font=("Poppins", 12)
                     ).grid(row=0, column=1, sticky="w", padx=(0, 16))

        tk.Label(ctrl, text="Room Filter:", fg=BLACK, bg=WHITE,
                 font=("Poppins", 12, "bold")).grid(row=0, column=2, sticky="w", padx=(0, 8))
        self.room_var = tk.StringVar(value="All Rooms")
        self.room_combo = ttk.Combobox(ctrl, textvariable=self.room_var,
                                       state="readonly", width=18,
                                       font=("Poppins", 12))
        self.room_combo.grid(row=0, column=3, sticky="w", padx=(0, 16))
        self._load_rooms()

        tk.Label(ctrl, text="From:", fg=BLACK, bg=WHITE,
                 font=("Poppins", 12)).grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.from_entry = make_date_entry(ctrl, default_date="2025-01-01")
        self.from_entry.grid(row=1, column=1, sticky="w", padx=(0, 16), pady=(8, 0))

        tk.Label(ctrl, text="To:", fg=BLACK, bg=WHITE,
                 font=("Poppins", 12)).grid(row=1, column=2, sticky="w", pady=(8, 0))
        today = date.today()
        self.to_entry = make_date_entry(ctrl, default_date=today.strftime("%Y-%m-%d"))
        self.to_entry.grid(row=1, column=3, sticky="w", padx=(0, 16), pady=(8, 0))

        tk.Button(ctrl, text="Generate Report",
                  fg=WHITE, bg=MAROON,
                  font=("Poppins", 13, "bold"),
                  relief="flat", bd=0, padx=22, pady=9, cursor="hand2",
                  command=self._generate
                  ).grid(row=1, column=4, sticky="w", pady=(8, 0))

        self._dl_btn = tk.Button(
            ctrl, text="Download Report",
            fg=MAROON, bg=GOLD,
            font=("Poppins", 13, "bold"),
            relief="flat", bd=0, padx=22, pady=9, cursor="hand2",
            state="disabled",
            command=self._open_download_popup,
        )
        self._dl_btn.grid(row=1, column=5, sticky="w", padx=(10, 0), pady=(8, 0))

        ttk.Separator(content, orient="horizontal"
                      ).grid(row=2, column=0, sticky="ew", pady=(0, 8))

        result_frame = tk.Frame(content, bg=LIGHT,
                                highlightbackground="#DDDDDD", highlightthickness=1)
        result_frame.grid(row=3, column=0, sticky="nsew")
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)

        self.result_text = tk.Text(
            result_frame,
            bg=LIGHT, fg=BLACK,
            font=("Courier New", 12),
            wrap="word",
            relief="flat", bd=0,
            padx=16, pady=12,
            state="disabled"
        )
        vsb = ttk.Scrollbar(result_frame, orient="vertical",
                            command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=vsb.set)
        self.result_text.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

    # ------------------------------------------------------------------
    def _load_rooms(self):
        try:
            rows = execute_query("SELECT room_number FROM Rooms ORDER BY room_number",
                                 fetch=True)
            values = ["All Rooms"] + [r["room_number"] for r in rows]
            self.room_combo["values"] = values
        except Exception:
            self.room_combo["values"] = ["All Rooms"]

    def _set_result(self, text):
        self.result_text.config(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("end", text)
        self.result_text.config(state="disabled")

    def _store_report(self, rtype, from_date, to_date, rows):
        self.current_report_data = rows
        self.current_report_type = rtype
        self.current_date_range  = (from_date, to_date)
        self._dl_btn.config(state="normal")

    def _get_header_info(self):
        from_date, to_date = self.current_date_range
        name = self.user_info.get("name", "Unknown")
        role = self.user_info.get("role", "").capitalize()
        uid  = self.user_info.get("user_id", "")
        now  = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        return {
            "title":         self.current_report_type.upper(),
            "downloaded_by": f"{name} [{role}]",
            "user_id":       str(uid),
            "generated_on":  now,
            "date_range":    f"{from_date} to {to_date}",
        }

    # ------------------------------------------------------------------
    def _generate(self):
        rtype     = self.rtype_var.get()
        room_filt = self.room_var.get()
        from_date = self.from_entry.get().strip()
        to_date   = self.to_entry.get().strip()

        self._dl_btn.config(state="disabled")
        self.current_report_data = []

        try:
            if rtype == "Room Usage Report":
                self._room_usage(room_filt, from_date, to_date)
            elif rtype == "Student Usage & Penalty Report":
                self._student_usage(from_date, to_date)
            elif rtype == "Violations Report":
                self._violations_report(room_filt, from_date, to_date)
        except Exception as exc:
            messagebox.showerror("Report Error", str(exc))

    # ------------------------------------------------------------------
    def _room_usage(self, room_filt, from_date, to_date):
        params = [from_date, to_date]
        extra  = ""
        if room_filt != "All Rooms":
            extra = " AND r.room_number = %s"
            params.append(room_filt)

        rows = execute_query(
            f"""SELECT r.room_number,  r.room_category,
                       COUNT(res.reservation_id)                                    AS total_reservations,
                       SUM(CASE WHEN res.status = 'checked_in'  THEN 1 ELSE 0 END) AS checkins,
                       SUM(CASE WHEN res.status = 'cancelled'   THEN 1 ELSE 0 END) AS cancellations,
                       SUM(CASE WHEN res.status = 'no_show'     THEN 1 ELSE 0 END) AS no_shows,
                       SUM(CASE WHEN res.status = 'reserved'    THEN 1 ELSE 0 END) AS reserved,
                       SUM(CASE WHEN res.status = 'completed'   THEN 1 ELSE 0 END) AS completed
                FROM Rooms r
                LEFT JOIN Reservations res
                       ON res.room_id = r.room_id
                      AND res.reservation_date BETWEEN %s AND %s
                WHERE 1=1{extra}
                GROUP BY r.room_id, r.room_number, r.room_category
                ORDER BY r.room_category, r.room_number""",
            params, fetch=True
        )
        if not rows:
            self._set_result("No data found for the selected range.")
            return

        W = 100
        lines = [
            f"ROOM USAGE REPORT  |  {from_date}  to  {to_date}",
            "=" * W,
            f"{'Room':<10}{'Category':<22}{'Total':>7}"
            f"{'Check-Ins':>11}{'Cancels':>10}{'No-Shows':>10}"
            f"{'Reserved':>10}{'Completed':>11}",
            "-" * W,
        ]
        for r in rows:
            lines.append(
                f"{r['room_number']:<10}{r['room_category']:<22}"
                f"{r['total_reservations']:>7}"
                f"{r['checkins']:>11}"
                f"{r['cancellations']:>10}"
                f"{r['no_shows']:>10}"
                f"{r['reserved']:>10}"
                f"{r['completed']:>11}"
            )
        lines.append("=" * W)
        self._set_result("\n".join(lines))
        self._store_report("Room Usage Report", from_date, to_date, rows)

    def _student_usage(self, from_date, to_date):
        rows = execute_query(
            """SELECT u.user_id, u.first_name, u.last_name, u.email,
                      u.penalty_points, u.account_status,
                      COUNT(res.reservation_id)                                    AS total_reservations,
                      SUM(CASE WHEN res.status = 'no_show'   THEN 1 ELSE 0 END)   AS no_shows,
                      SUM(CASE WHEN res.status = 'cancelled' THEN 1 ELSE 0 END)   AS cancellations
               FROM Users u
               LEFT JOIN Reservations res
                      ON res.user_id = u.user_id
                     AND res.reservation_date BETWEEN %s AND %s
               WHERE u.role = 'student'
               GROUP BY u.user_id
               ORDER BY u.penalty_points DESC""",
            (from_date, to_date), fetch=True
        )
        if not rows:
            self._set_result("No student data found.")
            return

        W = 104
        lines = [
            f"STUDENT USAGE & PENALTY REPORT  |  {from_date}  to  {to_date}",
            "=" * W,
            f"{'ID':<8}{'Name':<24}{'Email':<28}{'Pts':>5}"
            f"{'Status':>12}{'Total':>7}{'No-Shows':>10}{'Cancels':>10}",
            "-" * W,
        ]
        for r in rows:
            name = f"{r['first_name']} {r['last_name']}"
            lines.append(
                f"{r['user_id']:<8}{name:<24}{r['email']:<28}"
                f"{r['penalty_points']:>5}"
                f"{r['account_status']:>12}"
                f"{r['total_reservations']:>7}"
                f"{r['no_shows']:>10}"
                f"{r['cancellations']:>10}"
            )
        lines.append("=" * W)
        self._set_result("\n".join(lines))
        self._store_report("Student Usage & Penalty Report", from_date, to_date, rows)

    def _violations_report(self, room_filt, from_date, to_date):
        params = [from_date, to_date]
        extra  = ""
        if room_filt != "All Rooms":
            extra = " AND r.room_number = %s"
            params.append(room_filt)

        rows = execute_query(
            f"""SELECT v.violation_id,
                       u.first_name, u.last_name,
                       r.room_number,
                       v.violation_type, v.points_assessed,
                       v.status, v.created_at
                FROM Violations v
                JOIN Users u          ON u.user_id = v.user_id
                JOIN Reservations res ON res.reservation_id = v.reservation_id
                JOIN Rooms r          ON r.room_id = res.room_id
                WHERE DATE(v.created_at) BETWEEN %s AND %s{extra}
                ORDER BY v.created_at DESC""",
            params, fetch=True
        )
        if not rows:
            self._set_result("No violations found for the selected range.")
            return

        total_pts = sum(r["points_assessed"] or 0 for r in rows)
        lines = [
            f"VIOLATIONS REPORT  |  {from_date}  to  {to_date}",
            "=" * 74,
            f"{'ID':<6}{'Student':<26}{'Room':<10}{'Type':<18}"
            f"{'Pts':>5}{'Status':>10}{'Date':>12}",
            "-" * 74,
        ]
        for r in rows:
            name = f"{r['first_name']} {r['last_name']}"
            lines.append(
                f"{r['violation_id']:<6}{name:<26}"
                f"{r['room_number']:<10}{r['violation_type']:<18}"
                f"{r['points_assessed']:>5}{r['status']:>10}"
                f"{str(r['created_at'])[:10]:>12}"
            )
        lines += [
            "=" * 74,
            f"Total violations: {len(rows)}   Total penalty points issued: {total_pts}",
        ]
        self._set_result("\n".join(lines))
        self._store_report("Violations Report", from_date, to_date, rows)

    # ------------------------------------------------------------------
    # Download popup
    # ------------------------------------------------------------------
    def _open_download_popup(self):
        from_date, to_date = self.current_date_range
        rtype = self.current_report_type

        popup = tk.Toplevel(self)
        popup.title("Download Report")
        popup.resizable(False, False)
        popup.configure(bg=WHITE)
        popup.grab_set()

        popup.update_idletasks()
        pw, ph = 380, 340
        sx = self.winfo_rootx() + (self.winfo_width()  - pw) // 2
        sy = self.winfo_rooty() + (self.winfo_height() - ph) // 2
        popup.geometry(f"{pw}x{ph}+{sx}+{sy}")

        tk.Label(popup, text="Download Report",
                 fg=MAROON, bg=WHITE,
                 font=("Poppins", 16, "bold")
                 ).pack(pady=(20, 4))

        tk.Label(popup, text=f"{rtype}  |  {from_date}  to  {to_date}",
                 fg="#555555", bg=WHITE,
                 font=("Poppins", 10),
                 wraplength=340
                 ).pack(pady=(0, 16))

        btn_frame = tk.Frame(popup, bg=WHITE)
        btn_frame.pack(pady=(0, 12))

        for label, cmd in [
            ("Text File (.txt)",   lambda: self._download_txt(popup)),
            ("CSV File (.csv)",    lambda: self._download_csv(popup)),
            ("Excel File (.xlsx)", lambda: self._download_xlsx(popup)),
            ("PDF File (.pdf)",    lambda: self._download_pdf(popup)),
        ]:
            tk.Button(
                btn_frame, text=label,
                fg=WHITE, bg=MAROON,
                font=("Poppins", 12, "bold"),
                relief="flat", bd=0, padx=18, pady=8, cursor="hand2",
                command=cmd,
            ).pack(fill="x", pady=4, padx=32)

        tk.Button(
            popup, text="Cancel",
            fg=MAROON, bg="#E0E0E0",
            font=("Poppins", 11),
            relief="flat", bd=0, padx=18, pady=6, cursor="hand2",
            command=popup.destroy,
        ).pack(pady=(0, 16))

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------
    def _build_filename(self, ext):
        stem = FILENAME_MAP.get(self.current_report_type, "Report")
        uid  = self.user_info.get("user_id", 0)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{stem}_{uid}_{ts}.{ext}"

    # ------------------------------------------------------------------
    # Download: TXT
    # ------------------------------------------------------------------
    def _download_txt(self, popup):
        default = self._build_filename("txt")
        path = filedialog.asksaveasfilename(
            parent=popup,
            title="Save Text Report",
            initialfile=default,
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if not path:
            return

        h   = self._get_header_info()
        sep = "=" * 60
        header_lines = [
            sep,
            h["title"],
            sep,
            f"{'Downloaded By':<14}: {h['downloaded_by']}",
            f"{'User ID':<14}: {h['user_id']}",
            f"{'Generated On':<14}: {h['generated_on']}",
            f"{'Date Range':<14}: {h['date_range']}",
            sep,
            "",
        ]
        report_body = self.result_text.get("1.0", "end-1c")

        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(header_lines))
                f.write("\n")
                f.write(report_body)
            popup.destroy()
            messagebox.showinfo("Download Complete", f"Report saved:\n{path}")
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc))

    # ------------------------------------------------------------------
    # Download: CSV
    # ------------------------------------------------------------------
    def _download_csv(self, popup):
        default = self._build_filename("csv")
        path = filedialog.asksaveasfilename(
            parent=popup,
            title="Save CSV Report",
            initialfile=default,
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return

        h        = self._get_header_info()
        headers  = HEADERS_MAP.get(self.current_report_type, [])
        row_keys = ROW_KEYS_MAP.get(self.current_report_type, [])

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([h["title"]])
                writer.writerow([])
                writer.writerow(["Downloaded By:", h["downloaded_by"]])
                writer.writerow(["User ID:",       h["user_id"]])
                writer.writerow(["Generated On:",  h["generated_on"]])
                writer.writerow(["Date Range:",    h["date_range"]])
                writer.writerow([])
                writer.writerow(headers)
                for row in self.current_report_data:
                    writer.writerow([str(row.get(k, "")) for k in row_keys])
            popup.destroy()
            messagebox.showinfo("Download Complete", f"Report saved:\n{path}")
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc))

    # ------------------------------------------------------------------
    # Download: XLSX
    # ------------------------------------------------------------------
    def _download_xlsx(self, popup):
        if not _XLSX:
            messagebox.showerror(
                "Missing Library",
                "openpyxl is not installed.\nRun: pip install openpyxl"
            )
            return
        default = self._build_filename("xlsx")
        path = filedialog.asksaveasfilename(
            parent=popup,
            title="Save Excel Report",
            initialfile=default,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return

        h          = self._get_header_info()
        headers    = HEADERS_MAP.get(self.current_report_type, [])
        row_keys   = ROW_KEYS_MAP.get(self.current_report_type, [])
        sheet_name = FILENAME_MAP.get(self.current_report_type, "Report")[:31]
        num_cols   = max(len(headers), 2)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            maroon_fill = PatternFill("solid", fgColor="5E1219")
            white_fill  = PatternFill("solid", fgColor="FFFFFF")
            gold_fill   = PatternFill("solid", fgColor="FFF8E7")

            # Row 1: merged title
            title_cell = ws.cell(row=1, column=1, value=h["title"])
            title_cell.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
            title_cell.fill      = maroon_fill
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            ws.row_dimensions[1].height = 30

            # Row 2: blank

            # Rows 3-6: metadata
            label_font = Font(bold=True, color="5E1219", name="Calibri", size=11)
            value_font = Font(name="Calibri", size=11)
            meta_rows = [
                ("Downloaded By:", h["downloaded_by"]),
                ("User ID:",       h["user_id"]),
                ("Generated On:",  h["generated_on"]),
                ("Date Range:",    h["date_range"]),
            ]
            for i, (label, value) in enumerate(meta_rows, start=3):
                lc = ws.cell(row=i, column=1, value=label)
                lc.font = label_font
                vc = ws.cell(row=i, column=2, value=value)
                vc.font = value_font

            # Row 7: blank

            # Row 8: column headers
            hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=8, column=col_idx, value=header)
                cell.fill      = maroon_fill
                cell.font      = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[8].height = 20

            # Rows 9+: data with alternating row colors
            data_font = Font(name="Calibri", size=11)
            for row_idx, row in enumerate(self.current_report_data, start=9):
                fill = white_fill if row_idx % 2 == 0 else gold_fill
                for col_idx, key in enumerate(row_keys, start=1):
                    val  = row.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=str(val) if val is not None else "")
                    cell.fill      = fill
                    cell.font      = data_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(path)
            popup.destroy()
            messagebox.showinfo("Download Complete", f"Report saved:\n{path}")
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc))

    # ------------------------------------------------------------------
    # Download: PDF
    # ------------------------------------------------------------------
    def _download_pdf(self, popup):
        if not _PDF:
            messagebox.showerror(
                "Missing Library",
                "reportlab is not installed.\nRun: pip install reportlab"
            )
            return
        default = self._build_filename("pdf")
        path = filedialog.asksaveasfilename(
            parent=popup,
            title="Save PDF Report",
            initialfile=default,
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not path:
            return

        h        = self._get_header_info()
        headers  = HEADERS_MAP.get(self.current_report_type, [])
        row_keys = ROW_KEYS_MAP.get(self.current_report_type, [])

        RL_MAROON     = HexColor("#5E1219")
        RL_LIGHT_GOLD = HexColor("#FFF8E7")

        try:
            doc = SimpleDocTemplate(
                path,
                pagesize=landscape(letter),
                leftMargin=0.5 * inch,
                rightMargin=0.5 * inch,
                topMargin=0.5 * inch,
                bottomMargin=0.5 * inch,
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "ReportTitle",
                parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.white,
                backColor=RL_MAROON,
                spaceAfter=4,
                spaceBefore=4,
                alignment=1,
                fontName="Helvetica-Bold",
            )
            meta_label_style = ParagraphStyle(
                "MetaLabel",
                parent=styles["Normal"],
                fontSize=10,
                textColor=RL_MAROON,
                fontName="Helvetica-Bold",
            )
            meta_value_style = ParagraphStyle(
                "MetaValue",
                parent=styles["Normal"],
                fontSize=10,
                fontName="Helvetica",
            )

            story = []

            # Title block
            story.append(Paragraph(h["title"], title_style))
            story.append(Spacer(1, 0.15 * inch))

            # Metadata table
            meta_data = [
                [Paragraph("Downloaded By:", meta_label_style), Paragraph(h["downloaded_by"], meta_value_style)],
                [Paragraph("User ID:",        meta_label_style), Paragraph(h["user_id"],        meta_value_style)],
                [Paragraph("Generated On:",   meta_label_style), Paragraph(h["generated_on"],   meta_value_style)],
                [Paragraph("Date Range:",     meta_label_style), Paragraph(h["date_range"],     meta_value_style)],
            ]
            meta_table = Table(meta_data, colWidths=[1.5 * inch, 5.5 * inch])
            meta_table.setStyle(TableStyle([
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 0.2 * inch))

            # Data table
            usable_width = landscape(letter)[0] - inch
            num_cols = len(headers)
            col_w = usable_width / num_cols if num_cols else usable_width

            table_data = [headers]
            for row in self.current_report_data:
                table_data.append([str(row.get(k, "") or "") for k in row_keys])

            data_table = Table(table_data, colWidths=[col_w] * num_cols, repeatRows=1)
            data_table.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0),  RL_MAROON),
                ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0),  10),
                ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE",      (0, 1), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, RL_LIGHT_GOLD]),
                ("GRID",          (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ]))
            story.append(data_table)

            doc.build(story)
            popup.destroy()
            messagebox.showinfo("Download Complete", f"Report saved:\n{path}")
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc))
